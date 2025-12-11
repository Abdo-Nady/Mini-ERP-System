import openpyxl
from django.http import HttpResponse
from customer.models import Customer
from sales_order.models import SalesOrder, SalesOrderLine
from products.models import Product
from datetime import date
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAdminUser, IsAuthenticated
from rest_framework.response import Response


# CSV FILE
@api_view(['GET'])
@permission_classes([IsAdminUser])
def export_full_report(request):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # --- Customers Sheet ---
    ws_cust = wb.create_sheet(title="Customers")
    ws_cust.append([
        "Customer Code",
        "Name",
        "Phone",
        "Email",
        "Address",
        "Opening Balance"
    ])

    for c in Customer.objects.all():
        ws_cust.append([
            c.customer_code or 'N/A',
            c.name,
            c.phone,
            c.email,
            c.address,
            str(c.opening_balance) if c.opening_balance else '0.00'
        ])

    # --- Products Sheet ---
    ws_prod = wb.create_sheet(title="Products")
    ws_prod.append([
        "SKU",
        "Product Name",
        "Category",
        "Cost Price",
        "Selling Price",
        "Stock",
        "Has Image"
    ])

    for p in Product.objects.all():
        ws_prod.append([
            p.sku,
            p.name,
            p.category,
            str(p.cost_price),
            str(p.selling_price),
            p.stock,
            'Yes' if p.image else 'No'
        ])

    # --- Sales Orders Sheet ---
    ws_sales = wb.create_sheet(title="Sales Orders")
    ws_sales.append([
        "Order Number",
        "Customer Code",
        "Customer Name",
        "Order Date",
        "Created By",
        "Status",
        "Total Amount"
    ])

    for order in SalesOrder.objects.select_related('customer', 'created_by').all():
        ws_sales.append([
            order.order_number,
            order.customer.customer_code or 'N/A',
            order.customer.name,
            order.order_date.strftime('%Y-%m-%d %H:%M:%S'),
            order.created_by.username,
            order.get_status_display(),  # Get human-readable status
            str(order.total_amount)
        ])

    # --- Sales Order Lines Sheet ---
    ws_lines = wb.create_sheet(title="Order Lines")
    ws_lines.append([
        "Order Number",
        "Customer Name",
        "Product SKU",
        "Product Name",
        "Quantity",
        "Price",
        "Line Total"
    ])

    for order in SalesOrder.objects.select_related('customer').prefetch_related('lines__product').all():
        for line in order.lines.all():
            ws_lines.append([
                order.order_number,
                order.customer.name,
                line.product.sku,
                line.product.name,
                line.qty,
                str(line.price) if line.price else '0.00',
                str(line.line_total)
            ])

    # --- Stock Movement Logs Sheet ---
    from sales_order.models import StockMovementLog

    ws_logs = wb.create_sheet(title="Stock Movements")
    ws_logs.append([
        "Product SKU",
        "Product Name",
        "Quantity Change",
        "Action",
        "User",
        "Timestamp"
    ])

    for log in StockMovementLog.objects.select_related('product', 'user').order_by('-timestamp')[:1000]:
        action = "Added" if log.qty > 0 else "Removed"
        ws_logs.append([
            log.product.sku,
            log.product.name,
            abs(log.qty),
            action,
            log.user.username if log.user else 'System',
            log.timestamp.strftime('%Y-%m-%d %H:%M:%S')
        ])

    # --- Prepare response ---
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=full_report.xlsx'
    wb.save(response)
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard(request):
    # total customers
    total_customers = Customer.objects.count()

    # total sales today
    total_sales_today = SalesOrder.objects.filter(order_date__date=date.today(),
                                                  status=SalesOrder.Status.CONFIRMED ).count()


    low_stock_products = Product.objects.filter(stock__lte=5).values('name', 'stock')

    return Response({
        "total_customers": total_customers,
        "total_sales_today": total_sales_today,
        "stock_running_low": list(low_stock_products)
    })

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def menu(request):
    user = request.user
    if user.is_staff:
        items = ["dashboard", "customers", "sales", "products", "reports"]
    else:
        items = ["customers", "sales"]

    return Response({"menu": items})